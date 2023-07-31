import openpyxl
from time import sleep
import time

print('Music Name: ', end='')
file_name = input()

excel_file = f'excel_scores/{file_name}.xlsx'
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook['Sheet']

STEP = 1
OCTAVE = 2
CHORD = 3
DURATION = 4
BPM = 100 # "Andante"

i = 2
while worksheet.cell(row=i, column=STEP).value:
    if worksheet.cell(row=i, column=STEP).value == 'R': # rest
        rest = worksheet.cell(row=i, column=DURATION).value * 60 / BPM
        start_time = time.time()
        sleep(rest);
        end_time = time.time()
        execution_time = end_time - start_time
        print("Execution Time:", execution_time, "seconds")
    elif worksheet.cell(row=i, column=CHORD).value: # chord
        while worksheet.cell(row=i, column=CHORD).value:
            print('Chord:', worksheet.cell(row=i, column=STEP).value) # press the key
            i += 1
        i -= 1
    else: # solo key
        print(worksheet.cell(row=i, column=STEP).value) # press the key
    i += 1
