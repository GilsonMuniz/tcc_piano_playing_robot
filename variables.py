import openpyxl

PRESSED = False
NOT_PRESSED = True

CROTCHET_DURATION = 3
BPM = 100
SAMPLE_TIME = 60 / (CROTCHET_DURATION * BPM)

HEADER_ROW = ['Sample',
    'C1', 'C#1', 'Db1', 'D1', 'D#1', 'Eb1', 'E1', 'Fb1', 'E#1', 'F1', 'F#1', 'Gb1', 'G1', 'G#1', 'Ab1', 'A1', 'A#1', 'Bb1', 'B1', 'Cb1', 'B#1',
    'C2', 'C#2', 'Db2', 'D2', 'D#2', 'Eb2', 'E2', 'Fb2', 'E#2', 'F2', 'F#2', 'Gb2', 'G2', 'G#2', 'Ab2', 'A2', 'A#2', 'Bb2', 'B2', 'Cb2', 'B#2',
    'C3', 'C#3', 'Db3', 'D3', 'D#3', 'Eb3', 'E3', 'Fb3', 'E#3', 'F3', 'F#3', 'Gb3', 'G3', 'G#3', 'Ab3', 'A3', 'A#3', 'Bb3', 'B3', 'Cb3', 'B#3',
    'C4', 'C#4', 'Db4', 'D4', 'D#4', 'Eb4', 'E4', 'Fb4', 'E#4', 'F4', 'F#4', 'Gb4', 'G4', 'G#4', 'Ab4', 'A4', 'A#4', 'Bb4', 'B4', 'Cb4', 'B#4',
    'C5', 'C#5', 'Db5', 'D5', 'D#5', 'Eb5', 'E5', 'Fb5', 'E#5', 'F5', 'F#5', 'Gb5', 'G5', 'G#5', 'Ab5', 'A5', 'A#5', 'Bb5', 'B5', 'Cb5', 'B#5',
    'C6', 'C#6', 'Db6', 'D6', 'D#6', 'Eb6', 'E6', 'Fb6', 'E#6', 'F6', 'F#6', 'Gb6', 'G6', 'G#6', 'Ab6', 'A6', 'A#6', 'Bb6', 'B6', 'Cb6', 'B#6',
    'Rest']

NOTES_PINS = {'C3': 3, 'D3': 5, 'E3': 7, 'F3': 11, 'G3': 13, 'A3': 15, 'B3': 12, 'C4': 16, 'D4': 18, 'E4': 22, 'F4': 24, 'G4': 26, 'A4': 32, 'B4': 36, 'C5': 38, 'D5': 40}

COLUMNS_LETTERS = []
for col in range(1, 131 + 1): COLUMNS_LETTERS.append(openpyxl.utils.get_column_letter(col))
notes_columns = {key: value for key, value in zip(COLUMNS_LETTERS, HEADER_ROW)}

print('Music Name: ', end='')
file_name = input()

excel_file = f'excel_scores/{file_name}.xlsx'
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook['Sheet']
music_size = worksheet.max_row